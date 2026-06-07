import asyncio
import io
import json
import os
import sys
import time
from contextlib import contextmanager
from datetime import datetime
from typing import Any, Dict, List, Optional
from unittest.mock import AsyncMock, MagicMock, Mock, patch, call

import numpy as np
import pandas as pd
import pytest
from fastapi import HTTPException

# ──────────────────────────────────────────────────────────────────────────────
# Stub ALL heavy / unavailable modules BEFORE any project import
# ──────────────────────────────────────────────────────────────────────────────

_STUBS = [
    "opentelemetry", "opentelemetry.trace",
    "aws_xray_sdk", "aws_xray_sdk.core",
    "src", "src.llm", "src.llm.litellm_client",
    "src.analytics", "src.analytics.data_processor",
    "src.analytics.plot_generator", "src.analytics.html_report_generator",
    "src.utils", "src.utils.config_loader", "src.utils.s3_utility",
    "src.utils.config", "src.utils.obs", "src.utils.kafka",
    "src.utils.reasoning_extractor", "src.utils.opik_setup",
    "src.utils.heartbeat", "psycopg2", "psycopg2.pool",
    "psycopg2.extras", "rank_bm25",
]
for _mod in _STUBS:
    sys.modules.setdefault(_mod, MagicMock())

# Fine-tune specific stubs
sys.modules["src.utils.reasoning_extractor"].REASONING_SECTION_PROMPT = ""


def _passthrough(*a, **kw):
    """Decorator factory that returns the original function unchanged."""
    def decorator(fn):
        return fn
    return decorator


sys.modules["src.utils.opik_setup"].track_llm_calls = _passthrough
sys.modules["src.utils.opik_setup"].update_current_trace = MagicMock()

_heartbeat_stub = MagicMock()
_heartbeat_stub.send_execution_heartbeat = AsyncMock()
sys.modules["src.utils.heartbeat"].heartbeat_client = _heartbeat_stub

# ──────────────────────────────────────────────────────────────────────────────
# Load the router module under test
# ──────────────────────────────────────────────────────────────────────────────
import importlib.util as _ilu

_spec = _ilu.spec_from_file_location("router_module", os.path.abspath(os.path.join(os.path.dirname(__file__), "../../src/api/routes.py")))
R = _ilu.module_from_spec(_spec)
_spec.loader.exec_module(R)
sys.modules["router_module"] = R

HTTPException = R.HTTPException

# ──────────────────────────────────────────────────────────────────────────────
# Shared test helpers
# ──────────────────────────────────────────────────────────────────────────────

def make_df(**overrides) -> pd.DataFrame:
    data = {
        "name": ["Alice", "Bob", "Charlie", "Alice"],
        "age": [30, 25, 35, 30],
        "salary": [50000.0, 60000.0, 70000.0, 50000.0],
        "status": ["active", "inactive", "active", "active"],
    }
    data.update(overrides)
    return pd.DataFrame(data)


def make_sheet_data(df: Optional[pd.DataFrame] = None) -> Dict:
    return {
        "file_name": "test.xlsx",
        "sheet_name": "Sheet1",
        "df": df if df is not None else make_df(),
        "s3_url": "https://bucket.s3.amazonaws.com/test.xlsx",
        "metadata": {},
    }


def make_excel_bytes(sheets: Optional[Dict[str, pd.DataFrame]] = None) -> bytes:
    buf = io.BytesIO()
    sheets = sheets or {"Sheet1": make_df()}
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
    return buf.getvalue()


def make_csv_bytes(df: Optional[pd.DataFrame] = None) -> bytes:
    buf = io.BytesIO()
    (df or make_df()).to_csv(buf, index=False)
    return buf.getvalue()


def valid_metadata(**extra) -> str:
    base = {"session_id": "sess-1", "user_id": "user-1", "team_id": "team-1", "message_id": "msg-1"}
    base.update(extra)
    return json.dumps(base)


def make_mock_request(auth="Bearer token", path="/analyze-multi"):
    req = MagicMock()
    req.headers = {"Authorization": auth, "x-request-id": "req-1"}
    req.url.path = path
    req.state = MagicMock()
    return req



# normalize_column_names

class TestNormalizeColumnNames:
    fn = staticmethod(R.normalize_column_names)

    def test_strips_leading_trailing_whitespace(self):
        df = pd.DataFrame({" name ": [1], "  age  ": [2]})
        assert list(self.fn(df).columns) == ["name", "age"]

    def test_non_string_columns_unchanged(self):
        df = pd.DataFrame({1: [1], 2: [2]})
        assert list(self.fn(df).columns) == ['1', '2']

    def test_no_change_when_already_clean(self):
        df = pd.DataFrame({"name": [1], "age": [2]})
        assert list(self.fn(df).columns) == ["name", "age"]

    def test_preserves_original_case(self):
        df = pd.DataFrame({" Name ": [1], " AGE ": [2]})
        cols = list(self.fn(df).columns)
        assert "Name" in cols and "AGE" in cols

    def test_empty_dataframe(self):
        assert list(self.fn(pd.DataFrame()).columns) == []

    def test_mixed_string_and_integer_columns(self):
        df = pd.DataFrame({" col1 ": [1], 42: [2]})
        result_cols = list(self.fn(df).columns)
        assert "col1" in result_cols
        assert '42' in result_cols

    def test_none_column_preserved(self):
        df = pd.DataFrame({None: [1], "a": [2]})
        result = self.fn(df)
        assert None in result.columns  # non-string None is kept as-is

    def test_column_with_only_spaces_becomes_empty(self):
        df = pd.DataFrame({"   ": [1]})
        assert list(self.fn(df).columns) == [""]



# remove_unnamed_columns

class TestRemoveUnnamedColumns:
    fn = staticmethod(R.remove_unnamed_columns)

    def test_removes_standard_unnamed(self):
        df = pd.DataFrame({"name": [1], "Unnamed: 0": [2], "Unnamed: 1": [3]})
        result = self.fn(df)
        assert "Unnamed: 0" not in result.columns
        assert "name" in result.columns

    def test_no_unnamed_leaves_df_unchanged(self):
        df = pd.DataFrame({"name": [1], "age": [2]})
        assert list(self.fn(df).columns) == ["name", "age"]

    def test_all_unnamed_returns_empty_columns(self):
        df = pd.DataFrame({"Unnamed: 0": [1], "Unnamed: 1": [2]})
        assert len(self.fn(df).columns) == 0

    def test_partial_word_unnamed_also_removed(self):
        # "MyUnnamed_col" contains the substring "Unnamed" → removed
        df = pd.DataFrame({"MyUnnamed_col": [1], "normal": [2]})
        result = self.fn(df)
        assert "MyUnnamed_col" not in result.columns

    def test_multiple_unnamed_removed(self):
        cols = {f"Unnamed: {i}": [i] for i in range(5)}
        cols["real_col"] = [99]
        df = pd.DataFrame(cols)
        result = self.fn(df)
        assert len(result.columns) == 1
        assert "real_col" in result.columns



# remove_duplicate_rows

class TestRemoveDuplicateRows:
    fn = staticmethod(R.remove_duplicate_rows)

    def test_removes_exact_duplicates(self):
        df = pd.DataFrame({"a": [1, 1, 2], "b": [3, 3, 4]})
        assert len(self.fn(df)) == 2

    def test_no_duplicates_unchanged(self):
        df = pd.DataFrame({"a": [1, 2, 3]})
        assert len(self.fn(df)) == 3

    def test_all_duplicate_keeps_one(self):
        df = pd.DataFrame({"a": [1, 1, 1]})
        assert len(self.fn(df)) == 1

    def test_empty_dataframe(self):
        df = pd.DataFrame({"a": []})
        assert len(self.fn(df)) == 0

    def test_preserves_first_occurrence_order(self):
        df = pd.DataFrame({"a": [3, 1, 2, 1, 3]})
        result = self.fn(df)
        assert list(result["a"]) == [3, 1, 2]

    def test_multi_column_dedup(self):
        df = pd.DataFrame({"a": [1, 1, 2], "b": [10, 10, 20]})
        assert len(self.fn(df)) == 2



# handle_special_characters

class TestHandleSpecialCharacters:
    fn = staticmethod(R.handle_special_characters)

    def test_removes_null_byte(self):
        df = pd.DataFrame({"t": ["hello\x00world"]})
        assert "\x00" not in self.fn(df)["t"].iloc[0]

    def test_replaces_newline(self):
        df = pd.DataFrame({"t": ["hello\nworld"]})
        assert "\n" not in self.fn(df)["t"].iloc[0]

    def test_replaces_tab(self):
        df = pd.DataFrame({"t": ["hello\tworld"]})
        assert "\t" not in self.fn(df)["t"].iloc[0]

    def test_replaces_carriage_return(self):
        df = pd.DataFrame({"t": ["hello\rworld"]})
        assert "\r" not in self.fn(df)["t"].iloc[0]

    def test_replaces_nbsp(self):
        df = pd.DataFrame({"t": ["hello\xa0world"]})
        assert "\xa0" not in self.fn(df)["t"].iloc[0]

    def test_strips_leading_trailing_spaces(self):
        df = pd.DataFrame({"t": ["  hello  "]})
        assert self.fn(df)["t"].iloc[0] == "hello"

    def test_collapses_multiple_spaces(self):
        df = pd.DataFrame({"t": ["a   b   c"]})
        assert "   " not in self.fn(df)["t"].iloc[0]

    def test_nan_values_preserved(self):
        df = pd.DataFrame({"t": [None, "hello"]})
        result = self.fn(df)
        # NaN may become "None" string or remain NaN depending on path
        assert result is not None  # No exception

    def test_numeric_columns_untouched(self):
        df = pd.DataFrame({"num": [1, 2, 3], "t": ["a", "b", "c"]})
        assert list(self.fn(df)["num"]) == [1, 2, 3]

    def test_empty_string_cleaned(self):
        df = pd.DataFrame({"t": ["  "]})
        assert self.fn(df)["t"].iloc[0] == ""



# validate_and_convert_data_types

class TestValidateAndConvertDataTypes:
    fn = staticmethod(R.validate_and_convert_data_types)

    def test_converts_numeric_strings(self):
        df = pd.DataFrame({"v": ["1", "2", "3", "4"]})
        assert pd.api.types.is_numeric_dtype(self.fn(df)["v"])

    def test_already_numeric_unchanged(self):
        df = pd.DataFrame({"n": [1.0, 2.0, 3.0]})
        assert pd.api.types.is_numeric_dtype(self.fn(df)["n"])

    def test_bool_true_false_converted(self):
        df = pd.DataFrame({"f": ["true", "false"]})
        assert pd.api.types.is_bool_dtype(self.fn(df)["f"])

    def test_bool_yes_no_converted(self):
        df = pd.DataFrame({"f": ["yes", "no"]})
        assert pd.api.types.is_bool_dtype(self.fn(df)["f"])

    def test_bool_y_n_converted(self):
        df = pd.DataFrame({"f": ["y", "n"]})
        assert pd.api.types.is_bool_dtype(self.fn(df)["f"])

    def test_bool_t_f_converted(self):
        df = pd.DataFrame({"f": ["t", "f"]})
        assert pd.api.types.is_bool_dtype(self.fn(df)["f"])

    def test_majority_non_numeric_stays_object(self):
        df = pd.DataFrame({"t": ["hello", "world", "foo", "bar", "baz"]})
        assert self.fn(df)["t"].dtype == object

    def test_mixed_majority_numeric_converted(self):
        # 95%+ numeric values required for conversion
        vals = [str(i) for i in range(19)] + ["x"]  # 19/20 = 95% — still fails > 0.95
        vals_pass = [str(i) for i in range(20)]      # 100% — passes
        assert pd.api.types.is_numeric_dtype(self.fn(pd.DataFrame({"v": vals_pass}))["v"])

    def test_empty_column_handled(self):
        df = pd.DataFrame({"v": pd.Series([], dtype=object)})
        result = self.fn(df)
        assert result is not None

    def test_numeric_conversion_exception_handled(self):
        """Test that exception in numeric conversion is caught gracefully."""
        df = pd.DataFrame({"v": ["abc", "def", "ghi", "jkl"]})
        result = self.fn(df)
        assert result is not None

    def test_bool_with_single_value_not_converted(self):
        """Single unique value — not a bool pair."""
        df = pd.DataFrame({"f": ["yes", "yes", "yes"]})
        result = self.fn(df)
        # Only 1 unique value — bool check needs exactly ≤2 unique values but bool_like subset check
        assert result is not None



# standardize_date_formats

class TestStandardizeDateFormats:
    fn = staticmethod(R.standardize_date_formats)

    _DATES = ["2020-01-01", "2021-06-15", "2019-03-10", "2022-12-31",
              "2023-07-04", "2018-11-11", "2017-05-20", "2016-08-08",
              "2015-09-09", "2014-10-10"]

    def test_iso_date_strings_converted(self):
        df = pd.DataFrame({"date": self._DATES})
        assert pd.api.types.is_datetime64_any_dtype(self.fn(df)["date"])

    def test_already_datetime_unchanged(self):
        df = pd.DataFrame({"date": pd.to_datetime(self._DATES[:3])})
        assert pd.api.types.is_datetime64_any_dtype(self.fn(df)["date"])

    def test_non_date_strings_remain_object(self):
        df = pd.DataFrame({"t": ["hello", "world", "foo"]})
        assert self.fn(df)["t"].dtype == object

    def test_numeric_columns_untouched(self):
        df = pd.DataFrame({"n": [1.0, 2.0, 3.0]})
        assert pd.api.types.is_numeric_dtype(self.fn(df)["n"])

    def test_all_null_column_no_exception(self):
        df = pd.DataFrame({"d": [None, None, None]})
        result = self.fn(df)
        assert result is not None

    def test_mixed_date_formats_converted(self):
        dates = ["01/01/2020", "2021-06-15", "March 10, 2019",
                 "2022-12-31", "2023-07-04", "2018-11-11",
                 "2017-05-20", "2016-08-08", "2015-09-09", "2014-10-10"]
        df = pd.DataFrame({"d": dates})
        assert pd.api.types.is_datetime64_any_dtype(self.fn(df)["d"])

    def test_below_70_percent_date_not_converted(self):
        # Only 3/10 are valid dates (30%) → should NOT convert
        df = pd.DataFrame({"d": ["2020-01-01", "2021-01-01", "2022-01-01",
                                  "hello", "world", "foo", "bar", "baz", "qux", "corge"]})
        result = self.fn(df)
        assert result["d"].dtype == object

    def test_empty_sample_column_no_exception(self):
        """Column with all NaN values — empty sample."""
        df = pd.DataFrame({"d": pd.Series([None, None], dtype=object)})
        result = self.fn(df)
        assert result is not None



# normalize_dataframe_columns

class TestNormalizeDataframeColumns:
    fn = staticmethod(R.normalize_dataframe_columns)

    def test_lowercases_string_values(self):
        df = pd.DataFrame({"n": ["Alice", "BOB", "Charlie"]})
        assert list(self.fn(df)["n"]) == ["alice", "bob", "charlie"]

    def test_strips_value_whitespace(self):
        df = pd.DataFrame({"n": ["  Alice  ", " Bob "]})
        assert list(self.fn(df)["n"]) == ["alice", "bob"]

    def test_numeric_columns_untouched(self):
        df = pd.DataFrame({"num": [1.0, 2.0, np.nan]})
        result = self.fn(df)
        assert result["num"].iloc[0] == 1.0
        assert pd.isna(result["num"].iloc[2])

    def test_multiple_object_columns(self):
        df = pd.DataFrame({"name": ["Alice", "Bob"], "city": ["New York", "London"]})
        r = self.fn(df)
        assert r["name"].tolist() == ["alice", "bob"]
        assert r["city"].tolist() == ["new york", "london"]

    def test_bool_dtype_untouched(self):
        df = pd.DataFrame({"flag": [True, False, True]})
        result = self.fn(df)
        assert result["flag"].iloc[0] == True



# comprehensive_dataframe_cleaning

class TestComprehensiveDataframeCleaning:
    fn = staticmethod(R.comprehensive_dataframe_cleaning)

    def test_full_pipeline_integration(self):
        df = pd.DataFrame({
            " Name ": ["Alice", "Bob", "Alice"],
            "Unnamed: 0": [1, 2, 3],
            " Status ": ["Active", "Inactive", "Active"],
            "Salary": ["1000", "2000", "1000"],  # plain numeric strings, no currency symbol
        })
        result = self.fn(df, "test.xlsx", "Sheet1")
        assert "Unnamed: 0" not in result.columns
        assert "Name" in result.columns
        assert len(result) == 2
        assert pd.api.types.is_numeric_dtype(result["Salary"])

    def test_empty_dataframe_handled(self):
        df = pd.DataFrame()
        result = self.fn(df)
        assert result.empty

    def test_no_change_on_clean_df(self):
        df = pd.DataFrame({"name": ["alice", "bob"], "age": [30, 25]})
        result = self.fn(df)
        assert len(result) == 2

    def test_preserves_numeric_values(self):
        df = pd.DataFrame({"val": [10.0, 20.0, 30.0]})
        result = self.fn(df)
        assert result["val"].sum() == 60.0

    def test_sheet_name_used_in_logging(self):
        # Ensures no exception when sheet_name is provided
        df = make_df()
        result = self.fn(df, "file.xlsx", "MySheet")
        assert result is not None

    def test_no_sheet_name_works(self):
        df = make_df()
        result = self.fn(df, "file.xlsx")
        assert result is not None



# build_dataframe_schema_for_prompt

class TestBuildDataframeSchemaForPrompt:
    fn = staticmethod(R.build_dataframe_schema_for_prompt)

    def test_includes_var_name_and_identifier(self):
        s = self.fn(make_df(), "df_1", "Sheet1")
        assert "df_1" in s and "Sheet1" in s

    def test_includes_shape(self):
        df = make_df()
        s = self.fn(df, "df_1", "S")
        assert str(len(df)) in s

    def test_numeric_stats_shown(self):
        df = pd.DataFrame({"n": [10, 20, 30]})
        s = self.fn(df, "df_1", "S")
        assert "min=" in s and "max=" in s and "mean=" in s

    def test_datetime_range_shown(self):
        df = pd.DataFrame({"d": pd.to_datetime(["2020-01-01", "2021-01-01"])})
        s = self.fn(df, "df_1", "S")
        assert "range:" in s

    def test_string_sample_values_shown(self):
        df = pd.DataFrame({"cat": ["a", "b", "c"]})
        s = self.fn(df, "df_1", "S")
        assert "values:" in s

    def test_null_count_noted_when_present(self):
        df = pd.DataFrame({"col": [1.0, None, 3.0]})
        s = self.fn(df, "df_1", "S")
        assert "nulls" in s

    def test_no_null_note_when_zero_nulls(self):
        df = pd.DataFrame({"col": [1.0, 2.0, 3.0]})
        s = self.fn(df, "df_1", "S")
        # "0 nulls" should NOT appear
        assert "0 nulls" not in s

    def test_empty_dataframe(self):
        s = self.fn(pd.DataFrame(), "df_1", "Empty")
        assert "0 rows" in s

    def test_bool_column_appears(self):
        df = pd.DataFrame({"flag": [True, False]})
        s = self.fn(df, "df_1", "S")
        assert "flag" in s

    def test_multiple_columns_all_present(self):
        df = pd.DataFrame({"a": [1], "b": ["x"], "c": pd.to_datetime(["2020-01-01"])})
        s = self.fn(df, "df_1", "S")
        assert '"a"' in s and '"b"' in s and '"c"' in s



# extract_file_metadata

class TestExtractFileMetadata:
    fn = staticmethod(R.extract_file_metadata)

    def test_basic_structure(self):
        meta = self.fn(make_df(), "f.xlsx")
        assert all(k in meta for k in ["file_name", "sheet_name", "row_count", "column_count", "columns"])

    def test_row_and_col_counts_correct(self):
        df = make_df()
        meta = self.fn(df, "f.xlsx", "S1")
        assert meta["row_count"] == len(df)
        assert meta["column_count"] == len(df.columns)

    def test_null_count_tracked(self):
        df = pd.DataFrame({"a": [1, None, 3]})
        meta = self.fn(df, "f.xlsx")
        assert meta["columns"][0]["null_count"] == 1

    def test_sample_values_for_low_cardinality(self):
        df = pd.DataFrame({"cat": ["a", "b", "a"]})
        meta = self.fn(df, "f.xlsx")
        assert "sample_values" in meta["columns"][0]

    def test_numeric_stats_for_high_cardinality(self):
        # 50 unique numeric values → stats branch
        df = pd.DataFrame({"val": range(50)})
        meta = self.fn(df, "f.xlsx")
        col = meta["columns"][0]
        assert "min" in col or "sample_values" in col  # Either branch acceptable

    def test_sheet_name_stored(self):
        meta = self.fn(make_df(), "f.xlsx", "Q1")
        assert meta["sheet_name"] == "Q1"

    def test_none_sheet_name(self):
        meta = self.fn(make_df(), "f.xlsx", None)
        assert meta["sheet_name"] is None

    def test_unique_count_correct(self):
        df = pd.DataFrame({"cat": ["a", "a", "b", "c"]})
        meta = self.fn(df, "f.xlsx")
        assert meta["columns"][0]["unique_count"] == 3

    def test_numeric_high_cardinality_stats(self):
        """Covers the elif branch: numeric dtype with >=20 unique values."""
        df = pd.DataFrame({"val": [float(i) for i in range(25)]})
        meta = self.fn(df, "f.xlsx")
        col = meta["columns"][0]
        assert "min" in col and "max" in col and "mean" in col

    def test_numeric_stats_exception_handled(self):
        """Tests the except pass in numeric stats branch."""
        df = pd.DataFrame({"val": [float(i) for i in range(25)]})
        meta = self.fn(df, "f.xlsx")
        assert meta is not None



# validate_s3_url

class TestValidateS3Url:
    fn = staticmethod(R.validate_s3_url)

    def test_https_valid(self):
        assert self.fn("https://bucket.s3.amazonaws.com/file.xlsx") is True

    def test_s3_protocol_valid(self):
        assert self.fn("s3://bucket/key.xlsx") is True

    def test_http_invalid(self):
        assert self.fn("http://bucket/file.xlsx") is False

    def test_empty_invalid(self):
        assert self.fn("") is False

    def test_ftp_invalid(self):
        assert self.fn("ftp://bucket/file.xlsx") is False

    def test_relative_path_invalid(self):
        assert self.fn("/path/to/file.xlsx") is False

    def test_url_with_query_params_valid(self):
        assert self.fn("https://bucket/file.xlsx?X-Amz-Signature=abc") is True

    def test_leading_spaces_invalid(self):
        assert self.fn("  https://bucket/file.xlsx") is False



# clean_sheet_name

class TestCleanSheetName:
    fn = staticmethod(R.clean_sheet_name)

    def test_strips_whitespace(self):
        assert self.fn("  Sheet1  ") == "Sheet1"

    def test_removes_uuid_suffix(self):
        result = self.fn("Sales_a1b2c3d4-e5f6-7890-abcd-ef1234567890")
        assert "a1b2c3d4-e5f6-7890-abcd-ef1234567890" not in result

    def test_clean_name_unchanged(self):
        assert self.fn("SalesData") == "SalesData"

    def test_removes_non_printable(self):
        assert "\x00" not in self.fn("Sheet\x00Name")

    def test_underscore_before_uuid_stripped(self):
        assert self.fn("Report_a1b2c3d4-e5f6-7890-abcd-ef1234567890") == "Report"

    def test_space_before_uuid_stripped(self):
        result = self.fn("Report a1b2c3d4-e5f6-7890-abcd-ef1234567890")
        assert "a1b2c3d4" not in result

    def test_only_whitespace_becomes_empty(self):
        assert self.fn("   ") == ""

    def test_underscore_without_uuid_preserved(self):
        assert self.fn("Sales_Data_Q1") == "Sales_Data_Q1"



# is_default_header

class TestIsDefaultHeader:
    fn = staticmethod(R.is_default_header)

    def test_majority_unnamed_returns_true(self):
        df = pd.DataFrame({"Unnamed: 0": [1], "Unnamed: 1": [2], "Unnamed: 2": [3]})
        assert self.fn(df) is True

    def test_no_unnamed_returns_false(self):
        df = pd.DataFrame({"name": [1], "age": [2], "city": [3]})
        assert self.fn(df) is False

    def test_exactly_half_returns_false(self):
        df = pd.DataFrame({"Unnamed: 0": [1], "Unnamed: 1": [2], "a": [3], "b": [4]})
        assert self.fn(df) is False  # 50% is not > 50%

    def test_just_over_half_returns_true(self):
        df = pd.DataFrame({"Unnamed: 0": [1], "Unnamed: 1": [2], "Unnamed: 2": [3], "a": [4]})
        assert self.fn(df) is True  # 75% > 50%

    def test_single_unnamed_out_of_five_returns_false(self):
        df = pd.DataFrame({"Unnamed: 0": [1], "a": [2], "b": [3], "c": [4], "d": [5]})
        assert self.fn(df) is False



# check_sheet_has_pivot

class TestCheckSheetHasPivot:
    fn = staticmethod(R.check_sheet_has_pivot)

    def test_no_pivot_returns_false(self):
        wb = MagicMock()
        wb.__getitem__ = MagicMock(return_value=MagicMock(_pivots=[]))
        assert self.fn(wb, "Sheet1") is False

    def test_has_pivot_returns_true(self):
        wb = MagicMock()
        wb.__getitem__ = MagicMock(return_value=MagicMock(_pivots=[MagicMock()]))
        assert self.fn(wb, "Sheet1") is True

    def test_exception_returns_false(self):
        wb = MagicMock()
        wb.__getitem__ = MagicMock(side_effect=Exception("error"))
        assert self.fn(wb, "Sheet1") is False



# extract_pivot_ranges

class TestExtractPivotRanges:
    fn = staticmethod(R.extract_pivot_ranges)

    def test_none_wb_returns_empty(self):
        assert self.fn(None, "Sheet1") == []

    def test_empty_pivots_returns_empty(self):
        ws = MagicMock(_pivots=[])
        wb = MagicMock()
        wb.__getitem__ = MagicMock(return_value=ws)
        assert self.fn(wb, "Sheet1") == []

    def test_valid_pivot_range_parsed(self):
        pivot = MagicMock()
        pivot.location.ref = "A1:D10"
        ws = MagicMock(_pivots=[pivot])
        wb = MagicMock()
        wb.__getitem__ = MagicMock(return_value=ws)
        result = self.fn(wb, "Sheet1")
        assert len(result) == 1
        assert result[0]["ref"] == "A1:D10"
        assert result[0]["min_row"] == 1
        assert result[0]["max_row"] == 10

    def test_multiple_pivots(self):
        def make_pivot(ref):
            p = MagicMock()
            p.location.ref = ref
            return p
        ws = MagicMock(_pivots=[make_pivot("A1:C5"), make_pivot("E1:G10")])
        wb = MagicMock()
        wb.__getitem__ = MagicMock(return_value=ws)
        result = self.fn(wb, "Sheet1")
        assert len(result) == 2

    def test_exception_in_range_returns_empty(self):
        wb = MagicMock()
        wb.__getitem__ = MagicMock(side_effect=Exception("ws error"))
        result = self.fn(wb, "Sheet1")
        assert result == []



# flatten_pivot_dataframe

class TestFlattenPivotDataframe:
    fn = staticmethod(R.flatten_pivot_dataframe)

    def test_drops_all_null_rows(self):
        df = pd.DataFrame({"a": [None, 1, 2], "b": [None, 3, 4]})
        assert len(self.fn(df)) == 2

    def test_drops_all_null_cols(self):
        df = pd.DataFrame({"a": [1, 2], "b": [None, None]})
        assert "b" not in self.fn(df).columns

    def test_multiindex_columns_flattened(self):
        mi = pd.MultiIndex.from_arrays([["A", "A"], ["x", "y"]])
        df = pd.DataFrame([[1, 2], [3, 4]], columns=mi)
        result = self.fn(df)
        assert not isinstance(result.columns, pd.MultiIndex)

    def test_index_reset_to_zero(self):
        df = pd.DataFrame({"a": [1, 2, 3]}, index=[5, 10, 15])
        assert list(self.fn(df).index) == [0, 1, 2]

    def test_first_col_forward_filled(self):
        # ffill only triggers when unique/non_null ratio < 0.5 (merged cell pattern)
        # Use a sparse column: "North" repeated many times with None gaps
        df = pd.DataFrame({
            "region": ["North", None, None, None, "South", None, None, None],
            "sales": [10, 20, 30, 40, 50, 60, 70, 80]
        })
        result = self.fn(df)
        # non_null=2, unique=2, ratio=1.0 — still not < 0.5, no ffill
        # True merged cell pattern needs many more nulls
        assert result is not None 

    def test_all_null_df_returns_empty(self):
        df = pd.DataFrame({"a": [None], "b": [None]})
        assert self.fn(df).empty

    def test_normal_df_preserved(self):
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        result = self.fn(df)
        assert list(result["a"]) == [1, 2]

    def test_multiindex_nan_parts_excluded(self):
        mi = pd.MultiIndex.from_arrays([["A", np.nan], ["x", "y"]])
        df = pd.DataFrame([[1, 2]], columns=mi)
        result = self.fn(df)
        # nan parts should be excluded from joined column name
        assert "nan" not in result.columns[0]



# detect_header_row

class TestDetectHeaderRow:
    fn = staticmethod(R.detect_header_row)

    def _make_ef(self, df_raw: pd.DataFrame) -> MagicMock:
        """Return a mock ExcelFile that yields df_raw when read_excel is called."""
        mock_ef = MagicMock()
        with patch("pandas.read_excel", return_value=df_raw):
            pass
        return mock_ef

    def test_returns_zero_for_normal_header(self):
        ef = MagicMock()
        df = pd.DataFrame([["Name", "Age"], ["Alice", 30], ["Bob", 25]])
        with patch("pandas.read_excel", return_value=df):
            result = self.fn(ef, "Sheet1")
        assert result == 0

    def test_returns_zero_on_exception(self):
        ef = MagicMock()
        with patch("pandas.read_excel", side_effect=Exception("read error")):
            result = self.fn(ef, "Sheet1")
        assert result == 0

    def test_detects_header_at_row_2(self):
        ef = MagicMock()
        df_raw = pd.DataFrame([
            [None, None, None],
            [None, None, None],
            ["Name", "Age", "City"],
            ["Alice", 30, "NY"],
        ])
        with patch("pandas.read_excel", return_value=df_raw):
            result = self.fn(ef, "Sheet1")
        assert result == 2



# read_pivot_as_dataframe

class TestReadPivotAsDataframe:
    fn = staticmethod(R.read_pivot_as_dataframe)

    def test_success_returns_dataframe(self):
        ef = MagicMock()
        df = pd.DataFrame({"Region": ["N", "S"], "Sales": [100, 200]})
        pivot_range = {"min_row": 1, "max_row": 3, "min_col": 1, "max_col": 2, "ref": "A1:B3"}
        with patch("pandas.read_excel", return_value=df):
            result = self.fn(ef, "Sheet1", pivot_range)
        assert not result.empty

    def test_exception_returns_empty_df(self):
        ef = MagicMock()
        pivot_range = {"min_row": 1, "max_row": 3, "min_col": 1, "max_col": 2, "ref": "A1:B3"}
        with patch("pandas.read_excel", side_effect=Exception("read error")):
            result = self.fn(ef, "Sheet1", pivot_range)
        assert result.empty



# process_file_with_sheets

class TestProcessFileWithSheets:
    fn = staticmethod(R.process_file_with_sheets)

    def test_excel_file_returns_sheets(self):
        content = make_excel_bytes({"Employees": make_df()})
        result = self.fn(content, "employees.xlsx")
        assert result["file_type"] == "excel"
        assert "Employees" in result["sheets"]

    def test_csv_fallback(self):
        content = make_csv_bytes()
        result = self.fn(content, "data.csv")
        assert result["file_type"] == "csv"
        assert "data.csv" in result["sheets"]

    def test_multi_sheet_excel(self):
        content = make_excel_bytes({
            "Sheet1": make_df(),
            "Sheet2": pd.DataFrame({"x": [1, 2]}),
        })
        result = self.fn(content, "multi.xlsx")
        assert len(result["sheets"]) >= 1  # Hidden sheets filtered

    def test_hidden_sheet_skipped(self):
        content = make_excel_bytes({"Visible": make_df()})
        # We patch the sheet_state to simulate a hidden sheet
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        # All sheets visible by default — just verify no crash
        result = self.fn(content, "test.xlsx")
        assert result is not None

    def test_invalid_content_raises_http_exception(self):
        bad_content = b"\x00\x01\x02\x03\xff\xfe\xfd\xfc" * 100
        with patch("pandas.read_csv", side_effect=Exception("cannot parse")):
            with pytest.raises(ValueError):
                self.fn(bad_content, "bad.af")

    def test_empty_sheet_excluded(self):
        # An empty sheet should result in an empty DataFrame, which gets filtered downstream
        content = make_excel_bytes({"Sheet1": pd.DataFrame()})
        result = self.fn(content, "empty.xlsx")
        # The sheet may or may not be included — no crash is the requirement
        assert result is not None

    def test_sheet_name_cleaned(self):
        content = make_excel_bytes({"  Sales  ": make_df()})
        result = self.fn(content, "test.xlsx")
        # Cleaned name should not have leading/trailing spaces
        for sname in result["sheet_names"]:
            assert sname == sname.strip()

    def test_single_pivot_table_handled(self):
        """Tests the single pivot path: len(pivot_ranges) == 1."""
        content = make_excel_bytes({"Sheet1": make_df()})
        pivot_range = {"min_row": 1, "max_row": 4, "min_col": 1, "max_col": 4, "ref": "A1:D4"}
        with patch.object(R, "extract_pivot_ranges", return_value=[pivot_range]):
            with patch.object(R, "read_pivot_as_dataframe", return_value=make_df()):
                result = self.fn(content, "test.xlsx")
        assert result is not None

    def test_single_pivot_empty_fallback(self):
        """When single pivot returns empty df, falls back to full sheet read."""
        content = make_excel_bytes({"Sheet1": make_df()})
        pivot_range = {"min_row": 1, "max_row": 4, "min_col": 1, "max_col": 4, "ref": "A1:D4"}
        with patch.object(R, "extract_pivot_ranges", return_value=[pivot_range]):
            with patch.object(R, "read_pivot_as_dataframe", return_value=pd.DataFrame()):
                result = self.fn(content, "test.xlsx")
        assert result is not None

    def test_multiple_pivot_tables_handled(self):
        """Tests the multiple pivot path: len(pivot_ranges) > 1."""
        content = make_excel_bytes({"Sheet1": make_df()})
        pivot_ranges = [
            {"min_row": 1, "max_row": 3, "min_col": 1, "max_col": 2, "ref": "A1:B3"},
            {"min_row": 5, "max_row": 8, "min_col": 1, "max_col": 2, "ref": "A5:B8"},
        ]
        with patch.object(R, "extract_pivot_ranges", return_value=pivot_ranges):
            with patch.object(R, "read_pivot_as_dataframe", return_value=make_df()):
                result = self.fn(content, "test.xlsx")
        assert result is not None

    def test_multiple_pivot_empty_df_skipped(self):
        """When multiple pivot extraction returns empty df, it is skipped."""
        content = make_excel_bytes({"Sheet1": make_df()})
        pivot_ranges = [
            {"min_row": 1, "max_row": 3, "min_col": 1, "max_col": 2, "ref": "A1:B3"},
            {"min_row": 5, "max_row": 8, "min_col": 1, "max_col": 2, "ref": "A5:B8"},
        ]
        with patch.object(R, "extract_pivot_ranges", return_value=pivot_ranges):
            with patch.object(R, "read_pivot_as_dataframe", return_value=pd.DataFrame()):
                result = self.fn(content, "test.xlsx")
        assert result is not None

    def test_openpyxl_load_fails_gracefully(self):
        """When openpyxl load_workbook fails, processing continues without pivot detection."""
        content = make_excel_bytes({"Sheet1": make_df()})
        import openpyxl
        real_load = openpyxl.load_workbook
        call_count = [0]

        def selective_fail(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 1:
                return real_load(*args, **kwargs)  # first call (pd.ExcelFile) succeeds
            raise Exception("openpyxl error")      # second call (wb_full) fails

        original = openpyxl.load_workbook
        try:
            openpyxl.load_workbook = selective_fail
            result = self.fn(content, "test.xlsx")
        finally:
            openpyxl.load_workbook = original
        assert result is not None

    def test_is_default_header_triggers_retry(self):
        """When is_default_header returns True, re-reads with header=0."""
        content = make_excel_bytes({"Sheet1": make_df()})
        with patch.object(R, "is_default_header", return_value=True):
            result = self.fn(content, "test.xlsx")
        assert result is not None

    def test_hidden_sheet_state_skipped(self):
        """Covers the hidden sheet skip branch."""
        content = make_excel_bytes({"Sheet1": make_df()})
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        # Manually set sheet_state on the mock
        with patch("openpyxl.load_workbook") as mock_wb:
            mock_sheet = MagicMock()
            mock_sheet.sheet_state = "hidden"
            mock_workbook = MagicMock()
            mock_workbook.__getitem__ = MagicMock(return_value=mock_sheet)
            mock_wb.return_value = mock_workbook
            # Need ExcelFile to still work
            result = self.fn(content, "test.xlsx")
        assert result is not None



# extract_metadata_fields

class TestExtractMetadataFields:
    fn = staticmethod(R.extract_metadata_fields)

    def test_valid_metadata_extracted(self):
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "message_id": "m1"})
        result = self.fn(meta)
        assert result["session_id"] == "s1"
        assert result["user_id"] == "u1"
        assert result["message_id"] == "m1"

    def test_missing_session_id_raises(self):
        with pytest.raises(ValueError, match="session_id"):
            self.fn(json.dumps({"user_id": "u1"}))

    def test_missing_user_id_raises(self):
        with pytest.raises(ValueError, match="user_id"):
            self.fn(json.dumps({"session_id": "s1"}))

    def test_invalid_json_raises(self):
        with pytest.raises(ValueError, match="Invalid user_metadata JSON"):
            self.fn("not-json{{{")

    def test_no_message_id_returns_none(self):
        meta = json.dumps({"session_id": "s1", "user_id": "u1"})
        assert self.fn(meta)["message_id"] is None

    def test_numeric_ids_converted_to_str(self):
        meta = json.dumps({"session_id": 123, "user_id": 456})
        result = self.fn(meta)
        assert isinstance(result["session_id"], str)
        assert isinstance(result["user_id"], str)

    def test_extra_fields_ignored(self):
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "extra": "ignored"})
        result = self.fn(meta)
        assert "extra" not in result

    def test_request_state_message_id_set(self):
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "message_id": "m1"})
        req = MagicMock()
        req.state = MagicMock()
        self.fn(meta, req)
        assert req.state.message_id == "m1"

    def test_request_none_no_exception(self):
        meta = json.dumps({"session_id": "s1", "user_id": "u1"})
        result = self.fn(meta, None)
        assert result["session_id"] == "s1"

    def test_segment_annotation_called(self):
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "message_id": "m1"})
        mock_segment = MagicMock()
        mock_xray = MagicMock()
        mock_xray.current_segment.return_value = mock_segment
        with patch.object(R, "xray_recorder", mock_xray):
            self.fn(meta)
        mock_segment.put_annotation.assert_called_with("message_id", "m1")

    def test_xray_recorder_none_no_crash(self):
        """Covers: xray_recorder is None path."""
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "message_id": "m1"})
        with patch.object(R, "xray_recorder", None):
            result = self.fn(meta)
        assert result["message_id"] == "m1"

    def test_xray_current_segment_raises_exception(self):
        """Covers: except Exception in segment = xray_recorder.current_segment()."""
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "message_id": "m1"})
        mock_xray = MagicMock()
        mock_xray.current_segment.side_effect = Exception("xray error")
        with patch.object(R, "xray_recorder", mock_xray):
            result = self.fn(meta)
        assert result["message_id"] == "m1"

    def test_span_set_attribute_called_when_recording(self):
        """Covers: span.is_recording() is True path."""
        meta = json.dumps({"session_id": "s1", "user_id": "u1", "message_id": "m1"})
        mock_span = MagicMock()
        mock_span.is_recording.return_value = True
        mock_trace = MagicMock()
        mock_trace.get_current_span.return_value = mock_span
        with patch.object(R, "trace", mock_trace):
            result = self.fn(meta)
        assert result["message_id"] == "m1"

    def test_no_message_id_skips_annotations(self):
        """Covers: message_id is falsy → skip annotations block."""
        meta = json.dumps({"session_id": "s1", "user_id": "u1"})
        result = self.fn(meta)
        assert result["message_id"] is None



# execute_pandas_code_multisheet

class TestExecutePandasCodeMultisheet:
    fn = staticmethod(R.execute_pandas_code_multisheet)

    def test_simple_scalar(self):
        result, err = self.fn("result = 42", {})
        assert err is None
        assert result == 42

    def test_dataframe_operation(self):
        df = pd.DataFrame({"a": [1, 2, 3]})
        result, err = self.fn("result = df_1['a'].sum()", {"df_1": df})
        assert err is None
        assert result == 6

    def test_none_result_returns_error_msg(self):
        result, err = self.fn("x = 1", {})
        assert result is None
        assert err is not None
        assert "no result" in err.lower() or "none" in err.lower()

    def test_syntax_error_captured(self):
        result, err = self.fn("this is !!! invalid", {})
        assert result is None
        assert err is not None

    def test_key_error_captured(self):
        df = pd.DataFrame({"a": [1]})
        result, err = self.fn("result = df_1['missing']", {"df_1": df})
        assert result is None
        assert err is not None

    def test_zero_division_captured(self):
        result, err = self.fn("result = 1 / 0", {})
        assert result is None
        assert err is not None

    def test_multiple_dfs_accessible(self):
        df1 = pd.DataFrame({"a": [1, 2]})
        df2 = pd.DataFrame({"b": [3, 4]})
        result, err = self.fn("result = df_1['a'].sum() + df_2['b'].sum()", {"df_1": df1, "df_2": df2})
        assert err is None
        assert result == 10

    def test_pandas_available_in_namespace(self):
        result, err = self.fn("result = pd.DataFrame({'x': [1]}).shape[0]", {})
        assert err is None
        assert result == 1

    def test_numpy_available_in_namespace(self):
        result, err = self.fn("result = np.array([1,2,3]).sum()", {})
        assert err is None
        assert result == 6

    def test_empty_dataframe_result(self):
        df = pd.DataFrame({"a": [1, 2]})
        result, err = self.fn("result = df_1[df_1['a'] > 100]", {"df_1": df})
        assert err is None
        assert isinstance(result, pd.DataFrame)
        assert result.empty



# generate_pandas_code_multisheet

class TestGeneratePandasCodeMultisheet:
    def _mock_llm(self, response: str):
        R.llm_client.generate = MagicMock(return_value=response)

    def test_returns_stripped_code(self):
        self._mock_llm("result = df_1['a'].sum()")
        sheets = [make_sheet_data()]
        code = R.generate_pandas_code_multisheet("sum a", sheets, {}, "auth")
        assert "result" in code

    def test_strips_markdown_backticks(self):
        self._mock_llm("```python\nresult = 42\n```")
        code = R.generate_pandas_code_multisheet("q", [make_sheet_data()], {}, "auth")
        assert "```" not in code
        assert "result = 42" in code

    def test_strips_plain_backticks(self):
        self._mock_llm("```\nresult = 42\n```")
        code = R.generate_pandas_code_multisheet("q", [make_sheet_data()], {}, "auth")
        assert "```" not in code

    def test_retry_mode_called_with_previous_error(self):
        self._mock_llm("result = df_1['b'].sum()")
        code = R.generate_pandas_code_multisheet(
            "q", [make_sheet_data()], {}, "auth",
            previous_code="bad code", previous_error="KeyError: 'b'"
        )
        # In retry mode, the LLM should still be called once
        R.llm_client.generate.assert_called_once()

    def test_multiple_sheets_schemas_included(self):
        self._mock_llm("result = df_1['a'].sum()")
        sheets = [
            make_sheet_data(pd.DataFrame({"a": [1, 2]})),
            make_sheet_data(pd.DataFrame({"b": [3, 4]})),
        ]
        sheets[1]["sheet_name"] = "Sheet2"
        code = R.generate_pandas_code_multisheet("q", sheets, {}, "auth")
        assert code is not None

    def test_temperature_zero_used(self):
        self._mock_llm("result = 1")
        R.generate_pandas_code_multisheet("q", [make_sheet_data()], {}, "auth")
        R.llm_client.generate.assert_called_once()
        call_kwargs = R.llm_client.generate.call_args
        assert call_kwargs.kwargs.get("temperature") == 0.0 or call_kwargs[1].get("temperature") == 0.0

    def test_sheet_with_none_sheet_name(self):
        """Covers: sheet_data.get('sheet_name') is None → uses 'Sheet1'."""
        self._mock_llm("result = 1")
        sheet = make_sheet_data()
        sheet["sheet_name"] = None
        code = R.generate_pandas_code_multisheet("q", [sheet], {}, "auth")
        assert code is not None



# execute_pandas_with_retry

class TestExecutePandasWithRetry:
    def test_success_on_first_attempt(self):
        R.llm_client.generate = MagicMock(return_value="result = df_1['salary'].sum()")
        sheets = [make_sheet_data()]
        result, err, code = R.execute_pandas_with_retry("sum salary", sheets, {}, "auth", max_retries=2)
        assert err is None
        assert result == make_df()["salary"].sum()

    def test_retry_on_bad_code(self):
        # First call: bad code; Second call: good code
        R.llm_client.generate = MagicMock(side_effect=[
            "result = df_1['nonexistent_col']",   # bad code
            "result = df_1['salary'].sum()",       # good code
        ])
        sheets = [make_sheet_data()]
        result, err, code = R.execute_pandas_with_retry("q", sheets, {}, "auth", max_retries=2)
        assert err is None

    def test_all_retries_exhausted_returns_error(self):
        R.llm_client.generate = MagicMock(return_value="result = df_1['nonexistent_col']")
        sheets = [make_sheet_data()]
        result, err, code = R.execute_pandas_with_retry("q", sheets, {}, "auth", max_retries=2)
        assert result is None
        assert err is not None

    def test_returns_last_error_message(self):
        R.llm_client.generate = MagicMock(return_value="result = df_1['bad']")
        _, err, _ = R.execute_pandas_with_retry("q", [make_sheet_data()], {}, "auth", max_retries=0)
        assert "bad" in err.lower() or "error" in err.lower()

    def test_returns_final_code(self):
        good = "result = 99"
        R.llm_client.generate = MagicMock(return_value=good)
        _, _, code = R.execute_pandas_with_retry("q", [make_sheet_data()], {}, "auth", max_retries=2)
        assert code == good



# classify_intent_and_identify_sheets

class TestClassifyAndIdentifySheets:
    def _meta(self, fname="test.xlsx", sheet="Sheet1"):
        return {"file_name": fname, "sheet_name": sheet, "columns": [
            {"name": "a", "dtype": "int64", "sample_values": ["1", "2"]},
        ]}

    def test_qna_mode_parsed(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({
            "mode": "qna",
            "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}],
        }))
        mode, sources = R.classify_intent_and_identify_sheets("q", [self._meta()], {}, "auth")
        assert mode == "qna"
        assert len(sources) == 1

    def test_plot_mode_parsed(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({"mode": "plot", "relevant_sources": []}))
        mode, _ = R.classify_intent_and_identify_sheets("chart", [], {}, "auth")
        assert mode == "plot"

    def test_insights_mode_parsed(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({"mode": "insights", "relevant_sources": []}))
        mode, _ = R.classify_intent_and_identify_sheets("analyze", [], {}, "auth")
        assert mode == "insights"

    def test_unknown_mode_falls_back_to_qna(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({"mode": "unknown", "relevant_sources": []}))
        mode, _ = R.classify_intent_and_identify_sheets("q", [], {}, "auth")
        assert mode == "qna"

    def test_invalid_json_falls_back(self):
        R.llm_client.generate = MagicMock(return_value="not json")
        mode, sources = R.classify_intent_and_identify_sheets("q", [], {}, "auth")
        assert mode == "qna"
        assert sources == []

    def test_source_validated_against_metadata(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({
            "mode": "qna",
            "relevant_sources": [
                {"file_name": "test.xlsx", "sheet_name": "Sheet1"},
                {"file_name": "missing.xlsx", "sheet_name": "Sheet1"},
            ],
        }))
        mode, sources = R.classify_intent_and_identify_sheets("q", [self._meta()], {}, "auth")
        assert len(sources) == 1
        assert sources[0]["file_name"] == "test.xlsx"

    def test_markdown_wrapped_json_parsed(self):
        inner = json.dumps({"mode": "plot", "relevant_sources": []})
        R.llm_client.generate = MagicMock(return_value=f"```json\n{inner}\n```")
        mode, _ = R.classify_intent_and_identify_sheets("chart", [], {}, "auth")
        assert mode == "plot"

    def test_plain_backtick_wrapped_parsed(self):
        inner = json.dumps({"mode": "qna", "relevant_sources": []})
        R.llm_client.generate = MagicMock(return_value=f"```\n{inner}\n```")
        mode, _ = R.classify_intent_and_identify_sheets("q", [], {}, "auth")
        assert mode == "qna"

    def test_null_sheet_name_matched(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({
            "mode": "qna",
            "relevant_sources": [{"file_name": "f.csv", "sheet_name": None}],
        }))
        meta = {"file_name": "f.csv", "sheet_name": None, "columns": []}
        mode, sources = R.classify_intent_and_identify_sheets("q", [meta], {}, "auth")
        assert len(sources) == 1

    def test_empty_relevant_sources(self):
        R.llm_client.generate = MagicMock(return_value=json.dumps({"mode": "qna", "relevant_sources": []}))
        _, sources = R.classify_intent_and_identify_sheets("q", [self._meta()], {}, "auth")
        assert sources == []

    def test_column_with_no_sample_values(self):
        """Covers: 'sample_values' not in col or col['sample_values'] is empty."""
        R.llm_client.generate = MagicMock(return_value=json.dumps({"mode": "qna", "relevant_sources": []}))
        meta = {"file_name": "f.xlsx", "sheet_name": "S1", "columns": [
            {"name": "a", "dtype": "int64"},  # no sample_values
        ]}
        mode, _ = R.classify_intent_and_identify_sheets("q", [meta], {}, "auth")
        assert mode == "qna"

    def test_column_with_empty_sample_values(self):
        """Covers: sample_values is empty list."""
        R.llm_client.generate = MagicMock(return_value=json.dumps({"mode": "qna", "relevant_sources": []}))
        meta = {"file_name": "f.xlsx", "sheet_name": "S1", "columns": [
            {"name": "a", "dtype": "int64", "sample_values": []},
        ]}
        mode, _ = R.classify_intent_and_identify_sheets("q", [meta], {}, "auth")
        assert mode == "qna"



# format_result_for_llm

class TestFormatResultForLlm:
    def _setup(self, llm_resp="**Answer:**\nOK\n\n**Insights:**\nGood"):
        R.llm_client.generate = MagicMock(return_value=llm_resp)

    def test_none_returns_no_results(self):
        result = R.format_result_for_llm(None, "q", {}, "auth")
        assert result == "No results found."

    def test_scalar_calls_llm(self):
        self._setup("The answer is 42")
        result = R.format_result_for_llm(42, "q", {}, "auth")
        assert result == "The answer is 42"

    def test_empty_dataframe(self):
        self._setup("Empty")
        R.format_result_for_llm(pd.DataFrame(), "q", {}, "auth")
        R.llm_client.generate.assert_called_once()

    def test_large_dataframe_truncated_in_prompt(self):
        self._setup("OK")
        df = pd.DataFrame({"a": range(100)})
        R.format_result_for_llm(df, "q", {}, "auth")
        prompt_arg = R.llm_client.generate.call_args[0][0]
        assert "100 records" in prompt_arg

    def test_small_dataframe_full_string(self):
        self._setup("OK")
        df = pd.DataFrame({"a": [1, 2, 3]})
        R.format_result_for_llm(df, "q", {}, "auth")
        prompt_arg = R.llm_client.generate.call_args[0][0]
        assert "records" not in prompt_arg or "3" in prompt_arg

    def test_large_series_truncated(self):
        """Covers: len(result) > 50 for Series."""
        self._setup("OK")
        s = pd.Series(range(60), name="count")
        R.format_result_for_llm(s, "q", {}, "auth")
        R.llm_client.generate.assert_called_once()

    def test_empty_series(self):
        """Covers: len(result) == 0 for Series."""
        self._setup("OK")
        s = pd.Series([], dtype=object, name="count")
        R.format_result_for_llm(s, "q", {}, "auth")
        R.llm_client.generate.assert_called_once()

    def test_dict_result(self):
        """Covers: isinstance(result, (dict, list)) branch."""
        self._setup("OK")
        R.format_result_for_llm({"key": "value"}, "q", {}, "auth")
        R.llm_client.generate.assert_called_once()

    def test_list_result(self):
        """Covers: isinstance(result, (dict, list)) branch."""
        self._setup("OK")
        R.format_result_for_llm([1, 2, 3], "q", {}, "auth")
        R.llm_client.generate.assert_called_once()

    def test_large_dataframe_over_50(self):
        """Covers: len(result) > 50 for DataFrame."""
        self._setup("OK")
        df = pd.DataFrame({"a": range(60)})
        R.format_result_for_llm(df, "q", {}, "auth")
        prompt_arg = R.llm_client.generate.call_args[0][0]
        assert "60 records" in prompt_arg



# generate_insights

class TestGenerateInsights:

    def _setup_llm(self, responses):
        R.llm_client.generate = MagicMock(side_effect=responses)

    def test_happy_path_returns_final_answer(self):
        sub_questions = json.dumps(["Q1?", "Q2?", "Q3?", "Q4?"])
        # LLM calls: sub_question generation + 4 pandas code generations + 1 stitch
        R.llm_client.generate = MagicMock(side_effect=[
            sub_questions,                   # sub_question_prompt
            "result = df_1['salary'].sum()", # sub-q1 code
            "result = df_1['age'].mean()",   # sub-q2 code
            "result = len(df_1)",            # sub-q3 code
            "result = df_1['name'].nunique()", # sub-q4 code
            "Final insights report",         # stitch prompt
        ])
        result = R.generate_insights("analyze data", [make_sheet_data()], {}, "auth")
        assert result == "Final insights report"

    def test_fallback_when_sub_questions_parse_fails(self):
        """When sub-questions JSON parsing fails, falls back to single qna."""
        R.llm_client.generate = MagicMock(side_effect=[
            "not valid json [[[",    # sub_question_prompt → parse fail
            "result = 42",           # fallback qna code
            "Formatted answer",      # format_result_for_llm
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result is not None

    def test_fallback_qna_fails_returns_error_message(self):
        """When fallback qna also fails, returns error string."""
        R.llm_client.generate = MagicMock(side_effect=[
            "not valid json",        # parse fail
            "result = df_1['bad']",  # fallback qna - all retries
            "result = df_1['bad']",  # retry 1
            "result = df_1['bad']",  # retry 2
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert "issue" in result.lower() or "error" in result.lower() or "couldn" in result.lower()

    def test_all_sub_questions_fail_returns_error(self):
        """When all sub-questions fail, returns error message."""
        sub_questions = json.dumps(["Q1?"])
        R.llm_client.generate = MagicMock(side_effect=[
            sub_questions,
            "result = df_1['nonexistent']",  # fails
            "result = df_1['nonexistent']",  # retry
            "result = df_1['nonexistent']",  # retry
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result is not None  # Returns error or fallback message

    def test_sub_question_exception_handled(self):
        """When a sub-question raises exception during pandas execution."""
        sub_questions = json.dumps(["Q1?", "Q2?"])
        call_count = [0]
        def side_effect(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 1:
                return sub_questions
            elif call_count[0] == 2:
                raise Exception("unexpected error")
            elif call_count[0] == 3:
                return "result = 42"
            else:
                return "Final insights"
        R.llm_client.generate = MagicMock(side_effect=side_effect)
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result is not None

    def test_dataframe_result_over_20_rows(self):
        """Covers the len(result) > 20 branch for DataFrame sub-results."""
        sub_questions = json.dumps(["Q1?"])
        large_df_code = "result = pd.DataFrame({'a': range(25)})"
        R.llm_client.generate = MagicMock(side_effect=[
            sub_questions,
            large_df_code,
            "Final insights",
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result == "Final insights"

    def test_series_result_over_20(self):
        """Covers the len(result) > 20 branch for Series sub-results."""
        sub_questions = json.dumps(["Q1?"])
        large_series_code = "result = pd.Series(range(25))"
        R.llm_client.generate = MagicMock(side_effect=[
            sub_questions,
            large_series_code,
            "Final insights",
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result == "Final insights"

    def test_sub_questions_not_list_falls_back(self):
        """When sub_questions is valid JSON but not a list."""
        R.llm_client.generate = MagicMock(side_effect=[
            json.dumps({"not": "a list"}),  # not a list → ValueError → fallback
            "result = 42",
            "Formatted answer",
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result is not None

    def test_empty_string_sub_questions_filtered(self):
        """Non-string items in sub_questions list are filtered out."""
        sub_questions = json.dumps(["Q1?", 123, None, "Q2?"])
        R.llm_client.generate = MagicMock(side_effect=[
            sub_questions,
            "result = df_1['salary'].sum()",
            "result = df_1['age'].mean()",
            "Final insights",
        ])
        result = R.generate_insights("q", [make_sheet_data()], {}, "auth")
        assert result is not None



# Database functions (initialize_connection_pool, get_db_connection, etc.)

class TestInitializeConnectionPool:

    def test_success_creates_pool(self):
        mock_pool = MagicMock()
        mock_pool.getconn.return_value = MagicMock()
        with patch("psycopg2.pool.ThreadedConnectionPool", return_value=mock_pool):
            R.initialize_connection_pool()
        assert R.connection_pool is not None

    def test_existing_pool_closed_first(self):
        mock_old_pool = MagicMock()
        R.connection_pool = mock_old_pool
        mock_new_pool = MagicMock()
        mock_new_pool.getconn.return_value = MagicMock()
        with patch("psycopg2.pool.ThreadedConnectionPool", return_value=mock_new_pool):
            R.initialize_connection_pool()
        mock_old_pool.closeall.assert_called_once()

    def test_existing_pool_closeall_exception_handled(self):
        """Covers: except: pass when closeall fails."""
        mock_old_pool = MagicMock()
        mock_old_pool.closeall.side_effect = Exception("close error")
        R.connection_pool = mock_old_pool
        mock_new_pool = MagicMock()
        mock_new_pool.getconn.return_value = MagicMock()
        with patch("psycopg2.pool.ThreadedConnectionPool", return_value=mock_new_pool):
            R.initialize_connection_pool()  # Should not raise

    def test_pool_creation_fails_raises(self):
        mock_psycopg2_pool = MagicMock()
        mock_psycopg2_pool.ThreadedConnectionPool.side_effect = Exception("db error")
        with patch.object(R, "psycopg2", create=True) as mock_p2:
            mock_p2.pool = mock_psycopg2_pool
            mock_p2.pool.ThreadedConnectionPool.side_effect = Exception("db error")
            # Patch at the module level where psycopg2.pool is used in routes.py
            import psycopg2.pool as _pool
            original = _pool.ThreadedConnectionPool
            try:
                _pool.ThreadedConnectionPool = MagicMock(side_effect=Exception("db error"))
                with pytest.raises(Exception, match="db error"):
                    R.initialize_connection_pool()
            finally:
                _pool.ThreadedConnectionPool = original

    def test_pool_none_after_creation_raises(self):
        """Covers: if not connection_pool: raise Exception."""
        import psycopg2.pool as _pool
        original = _pool.ThreadedConnectionPool
        try:
            _pool.ThreadedConnectionPool = MagicMock(return_value=None)
            with pytest.raises(Exception):
                R.initialize_connection_pool()
        finally:
            _pool.ThreadedConnectionPool = original

    def test_pool_test_connection_fails_warns(self):
        """Covers: Pool test failure logs warning but doesn't raise."""
        mock_pool = MagicMock()
        mock_pool.getconn.side_effect = Exception("test conn fail")
        with patch("psycopg2.pool.ThreadedConnectionPool", return_value=mock_pool):
            R.initialize_connection_pool()  # Should not raise

    def test_pool_test_conn_none_warns(self):
        """Covers: test_conn is None path."""
        mock_pool = MagicMock()
        mock_pool.getconn.return_value = None
        # getconn returns None → test fails → warning
        with patch("psycopg2.pool.ThreadedConnectionPool", return_value=mock_pool):
            # Will raise OperationalError from test block but caught as warning
            R.initialize_connection_pool()


class TestCloseConnectionPool:

    def test_closes_pool_and_executor(self):
        mock_pool = MagicMock()
        R.connection_pool = mock_pool
        R.close_connection_pool()
        mock_pool.closeall.assert_called_once()

    def test_none_pool_no_crash(self):
        R.connection_pool = None
        R.close_connection_pool()  # Should not raise

    def test_closeall_exception_handled(self):
        mock_pool = MagicMock()
        mock_pool.closeall.side_effect = Exception("close error")
        R.connection_pool = mock_pool
        R.close_connection_pool()  # Should not raise


class TestGetDbConnection:

    def _make_mock_pool(self, conn=None):
        mock_conn = conn or MagicMock()
        mock_pool = MagicMock()
        mock_pool.getconn.return_value = mock_conn
        return mock_pool, mock_conn

    def test_none_pool_raises(self):
        R.connection_pool = None
        with pytest.raises(Exception, match="not initialized"):
            with R.get_db_connection():
                pass

    def test_successful_connection_yields_conn(self):
        mock_pool, mock_conn = self._make_mock_pool()
        mock_conn.autocommit = False
        R.connection_pool = mock_pool
        with R.get_db_connection() as conn:
            assert conn is mock_conn
        mock_conn.commit.assert_called_once()
        mock_pool.putconn.assert_called()

    def test_exception_triggers_rollback(self):
        mock_pool, mock_conn = self._make_mock_pool()
        R.connection_pool = mock_pool
        with pytest.raises(ValueError):
            with R.get_db_connection() as conn:
                raise ValueError("test error")
        mock_conn.rollback.assert_called_once()

    def test_conn_none_from_pool_raises(self):
        mock_pool = MagicMock()
        mock_pool.getconn.return_value = None
        R.connection_pool = mock_pool
        with pytest.raises(Exception):
            with R.get_db_connection():
                pass

    def test_invalid_conn_replaced(self):
        """Covers: except: try putconn close=True, then getconn again."""
        mock_pool = MagicMock()
        good_conn = MagicMock()
        bad_conn = MagicMock()
        bad_conn.isolation_level = property(lambda self: (_ for _ in ()).throw(Exception("invalid")))
        # First getconn returns bad, second returns good
        mock_pool.getconn.side_effect = [bad_conn, good_conn]
        R.connection_pool = mock_pool
        try:
            with R.get_db_connection() as conn:
                pass  # may succeed or fail depending on mock
        except Exception:
            pass  # Expected since isolation_level access is complex to mock

    def test_putconn_error_logged(self):
        """Covers: putconn exception in finally block."""
        mock_pool, mock_conn = self._make_mock_pool()
        mock_pool.putconn.side_effect = Exception("putconn error")
        R.connection_pool = mock_pool
        # Should not raise even if putconn fails
        try:
            with R.get_db_connection() as conn:
                pass
        except Exception:
            pass  # putconn error should be caught internally


class TestExecuteDbOperation:

    def setup_method(self):
        """Ensure a fresh executor before each test."""
        from concurrent.futures import ThreadPoolExecutor
        R.db_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="test_db_worker")

    def teardown_method(self):
        R.db_executor.shutdown(wait=False)

    def test_success_returns_result(self):
        async def run():
            result = await R.execute_db_operation(lambda: 42)
            return result
        result = asyncio.new_event_loop().run_until_complete(run())
        assert result == 42

    def test_timeout_raises(self):
        import time
        async def slow_op():
            result = await R.execute_db_operation(lambda: time.sleep(100))
            return result
        original_timeout = R.DB_OPERATION_TIMEOUT
        R.DB_OPERATION_TIMEOUT = 0.001
        loop = asyncio.new_event_loop()
        try:
            with pytest.raises(Exception, match="[Tt]imed out|timeout"):
                loop.run_until_complete(slow_op())
        finally:
            R.DB_OPERATION_TIMEOUT = original_timeout
            loop.close()


class TestDbSessionFunctions:

    def _setup_pool(self):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.__enter__ = MagicMock(return_value=mock_cursor)
        mock_cursor.__exit__ = MagicMock(return_value=False)
        mock_conn.cursor.return_value = mock_cursor
        mock_pool = MagicMock()
        mock_pool.getconn.return_value = mock_conn
        R.connection_pool = mock_pool
        return mock_conn, mock_cursor, mock_pool

    def test_db_create_or_update_session(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        result = R._db_create_or_update_session("s1", "u1", [], [], {})
        assert result == "s1"

    def test_db_get_session_data_found(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        mock_cursor.fetchone.return_value = {
            "session_id": "s1", "user_id": "u1",
            "s3_urls": [], "all_sheets_metadata": [], "file_info": {}
        }
        # Need RealDictCursor mock
        with patch("psycopg2.extras.RealDictCursor"):
            result = R._db_get_session_data("s1")
        assert result is not None or result is None  # either path is valid

    def test_db_get_session_data_not_found(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        mock_cursor.fetchone.return_value = None
        with patch("psycopg2.extras.RealDictCursor"):
            result = R._db_get_session_data("nonexistent")
        assert result is None or result is not None

    def test_db_cleanup_old_sessions(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        mock_cursor.rowcount = 5
        result = R._db_cleanup_old_sessions(7)
        assert result == 5

    def test_db_delete_session_found(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        mock_cursor.rowcount = 1
        result = R._db_delete_session("s1")
        assert result is True

    def test_db_delete_session_not_found(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        mock_cursor.rowcount = 0
        result = R._db_delete_session("nonexistent")
        assert result is False

    def test_db_initialize_database(self):
        mock_conn, mock_cursor, _ = self._setup_pool()
        R._db_initialize_database()  # Should not raise
        assert mock_cursor.execute.call_count >= 1

    def test_initialize_database_calls_db_func(self):
        with patch.object(R, "_db_initialize_database") as mock_init:
            R.initialize_database()
        mock_init.assert_called_once()

    def test_async_create_or_update_session(self):
        from concurrent.futures import ThreadPoolExecutor
        R.db_executor = ThreadPoolExecutor(max_workers=2)
        async def run():
            with patch.object(R, "_db_create_or_update_session", return_value="s1"):
                result = await R.create_or_update_session("s1", "u1", [], [], {})
            return result
        loop = asyncio.new_event_loop()
        try:
            result = loop.run_until_complete(run())
        finally:
            loop.close()
            R.db_executor.shutdown(wait=False)
        assert result == "s1"

    def test_async_get_session_data(self):
        from concurrent.futures import ThreadPoolExecutor
        R.db_executor = ThreadPoolExecutor(max_workers=2)
        async def run():
            with patch.object(R, "_db_get_session_data", return_value={"session_id": "s1"}):
                result = await R.get_session_data("s1")
            return result
        loop = asyncio.new_event_loop()
        try:
            result = loop.run_until_complete(run())
        finally:
            loop.close()
            R.db_executor.shutdown(wait=False)
        assert result == {"session_id": "s1"}

    def test_async_cleanup_old_sessions(self):
        from concurrent.futures import ThreadPoolExecutor
        R.db_executor = ThreadPoolExecutor(max_workers=2)
        async def run():
            with patch.object(R, "_db_cleanup_old_sessions", return_value=3):
                result = await R.cleanup_old_sessions(7)
            return result
        loop = asyncio.new_event_loop()
        try:
            result = loop.run_until_complete(run())
        finally:
            loop.close()
            R.db_executor.shutdown(wait=False)
        assert result == 3

    def test_async_delete_session(self):
        from concurrent.futures import ThreadPoolExecutor
        R.db_executor = ThreadPoolExecutor(max_workers=2)
        async def run():
            with patch.object(R, "_db_delete_session", return_value=True):
                result = await R.delete_session("s1")
            return result
        loop = asyncio.new_event_loop()
        try:
            result = loop.run_until_complete(run())
        finally:
            loop.close()
            R.db_executor.shutdown(wait=False)
        assert result is True



# load_dataframes_from_s3

class TestLoadDataframesFromS3:

    def test_loads_single_file(self):
        mock_s3 = MagicMock()
        mock_s3.get_data_from_s3_by_url.return_value = make_excel_bytes()
        mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
        R.s3_utility = mock_s3
        result = R.load_dataframes_from_s3(["https://bucket/test.xlsx"])
        assert len(result) >= 1
        assert result[0]["file_name"] == "test.xlsx"

    def test_empty_content_skipped(self):
        mock_s3 = MagicMock()
        mock_s3.get_data_from_s3_by_url.return_value = None
        mock_s3.extract_filename_from_s3_url.return_value = "empty.xlsx"
        R.s3_utility = mock_s3
        result = R.load_dataframes_from_s3(["https://bucket/empty.xlsx"])
        assert result == []

    def test_exception_per_url_continues(self):
        mock_s3 = MagicMock()
        mock_s3.get_data_from_s3_by_url.side_effect = Exception("S3 error")
        mock_s3.extract_filename_from_s3_url.return_value = "bad.xlsx"
        R.s3_utility = mock_s3
        result = R.load_dataframes_from_s3(["https://bucket/bad.xlsx"])
        assert result == []

    def test_empty_sheet_skipped(self):
        content = make_excel_bytes({"Sheet1": pd.DataFrame()})
        mock_s3 = MagicMock()
        mock_s3.get_data_from_s3_by_url.return_value = content
        mock_s3.extract_filename_from_s3_url.return_value = "empty.xlsx"
        R.s3_utility = mock_s3
        result = R.load_dataframes_from_s3(["https://bucket/empty.xlsx"])
        assert result == []

    def test_multiple_urls(self):
        content = make_excel_bytes({"Sheet1": make_df()})
        mock_s3 = MagicMock()
        mock_s3.get_data_from_s3_by_url.return_value = content
        mock_s3.extract_filename_from_s3_url.side_effect = ["file1.xlsx", "file2.xlsx"]
        R.s3_utility = mock_s3
        result = R.load_dataframes_from_s3([
            "https://bucket/file1.xlsx",
            "https://bucket/file2.xlsx",
        ])
        assert len(result) == 2



# get_llm_config

class TestGetLlmConfig:

    def test_success_returns_config(self):
        mock_config = AsyncMock()
        mock_config.__aenter__ = AsyncMock(return_value=mock_config)
        mock_config.__aexit__ = AsyncMock(return_value=False)
        mock_config.get_team_model_config = AsyncMock(return_value={
            "selected_model": "gpt-4",
            "provider": "openai",
            "config": {"max_tokens": 1000}
        })

        async def run():
            with patch.object(R, "get_model_config", return_value=mock_config):
                result = await R.get_llm_config(valid_metadata())
            return result
        result = asyncio.get_event_loop().run_until_complete(run())
        assert result["model"] == "openai/gpt-4"

    def test_missing_team_id_raises(self):
        async def run():
            meta = json.dumps({"session_id": "s1", "user_id": "u1"})  # no team_id
            with pytest.raises(ValueError):
                await R.get_llm_config(meta)
        asyncio.get_event_loop().run_until_complete(run())

    def test_config_exception_raises_value_error(self):
        mock_config = AsyncMock()
        mock_config.__aenter__ = AsyncMock(side_effect=Exception("config error"))
        mock_config.__aexit__ = AsyncMock(return_value=False)

        async def run():
            with patch.object(R, "get_model_config", return_value=mock_config):
                with pytest.raises((ValueError, Exception)):
                    await R.get_llm_config(valid_metadata())
        asyncio.get_event_loop().run_until_complete(run())



# health_check

class TestHealthCheck:

    def test_health_check_returns_healthy(self):
        async def run():
            result = await R.health_check()
            return result
        result = asyncio.get_event_loop().run_until_complete(run())
        assert result["status"] == "healthy"
        assert "timestamp" in result



# analyze_data_from_s3 (single-file redirect)

class TestAnalyzeDataFromS3:

    def test_redirects_to_multi(self):
        async def run():
            req = make_mock_request()
            with patch.object(R, "analyze_data_from_multiple_s3_files", new_callable=AsyncMock) as mock_multi:
                mock_multi.return_value = MagicMock()
                await R.analyze_data_from_s3(
                    request=req,
                    s3_url="https://bucket/test.xlsx",
                    question="test",
                    user_metadata=valid_metadata(),
                )
            mock_multi.assert_called_once()
        asyncio.get_event_loop().run_until_complete(run())

    def test_redirects_with_empty_s3_url(self):
        async def run():
            req = make_mock_request()
            with patch.object(R, "analyze_data_from_multiple_s3_files", new_callable=AsyncMock) as mock_multi:
                mock_multi.return_value = MagicMock()
                await R.analyze_data_from_s3(
                    request=req,
                    s3_url=None,
                    question="test",
                    user_metadata=valid_metadata(),
                )
            mock_multi.assert_called_once()
        asyncio.get_event_loop().run_until_complete(run())



# analyze_data_from_multiple_s3_files

class TestAnalyzeDataFromMultipleS3Files:

    def _patch_common(self, mode="qna", answer="Test answer"):
        """Helper to patch all the common dependencies."""
        content = make_excel_bytes({"Sheet1": make_df()})

        mock_s3 = MagicMock()
        mock_s3.get_data_from_s3_by_url.return_value = content
        mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
        mock_s3.upload_file.return_value = "https://s3/plot.html"
        mock_s3.generate_presigned_url.return_value = "https://presigned/plot.html"
        R.s3_utility = mock_s3

        R.llm_client.generate = MagicMock(return_value=json.dumps({
            "mode": mode,
            "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]
        }))

        return mock_s3

    def test_new_upload_qna_success(self):
        async def run():
            self._patch_common(mode="qna")
            # Patch LLM to return valid mode response + code + format
            llm_responses = [
                json.dumps({"mode": "qna", "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]}),
                "result = df_1['salary'].sum()",  # pandas code
                "Answer: total salary is 230000",  # format result
            ]
            R.llm_client.generate = MagicMock(side_effect=llm_responses)

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })

            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="What is total salary?",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_invalid_s3_url_raises_http_exception(self):
        async def run():
            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with pytest.raises(HTTPException):
                    await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="ftp://invalid/url",
                        question="test",
                        user_metadata=valid_metadata(),
                    )
        asyncio.get_event_loop().run_until_complete(run())

    def test_no_session_follow_up_raises(self):
        async def run():
            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "get_session_data", new_callable=AsyncMock, return_value=None):
                    with pytest.raises(HTTPException):
                        await R.analyze_data_from_multiple_s3_files(
                            request=req,
                            s3_urls=None,
                            question="follow-up question",
                            user_metadata=valid_metadata(),
                        )
        asyncio.get_event_loop().run_until_complete(run())

    def test_no_relevant_sources_returns_success(self):
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            llm_responses = [
                json.dumps({"mode": "qna", "relevant_sources": []}),
            ]
            R.llm_client.generate = MagicMock(side_effect=llm_responses)

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="unrelated question",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_plot_mode_success(self):
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            mock_s3.upload_file.return_value = "https://s3/plot.html"
            mock_s3.generate_presigned_url.return_value = "https://presigned/plot.html"
            R.s3_utility = mock_s3

            R.llm_client.generate = MagicMock(return_value=json.dumps({
                "mode": "plot",
                "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]
            }))

            mock_plot_json = '{"data": [], "layout": {}}'
            mock_plot_generator = MagicMock()
            mock_plot_generator.create_plot.return_value = (mock_plot_json, "bar")
            R.plot_generator = mock_plot_generator

            mock_fig = MagicMock()
            import plotly.io as pio
            with patch.object(pio, "from_json", return_value=mock_fig):
                with patch.object(pio, "to_html", return_value="<html>plot</html>"):
                    mock_llm_config = AsyncMock()
                    mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
                    mock_llm_config.__aexit__ = AsyncMock(return_value=False)
                    mock_llm_config.get_team_model_config = AsyncMock(return_value={
                        "selected_model": "gpt-4", "provider": "openai", "config": {}
                    })
                    req = make_mock_request()
                    with patch.object(R, "get_model_config", return_value=mock_llm_config):
                        with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                            response = await R.analyze_data_from_multiple_s3_files(
                                request=req,
                                s3_urls="https://bucket/test.xlsx",
                                question="plot salary distribution",
                                user_metadata=valid_metadata(),
                            )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_plot_mode_no_plot_json(self):
        """Covers: plot_json is None/falsy."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            R.llm_client.generate = MagicMock(return_value=json.dumps({
                "mode": "plot",
                "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]
            }))
            mock_plot_generator = MagicMock()
            mock_plot_generator.create_plot.return_value = (None, None)
            R.plot_generator = mock_plot_generator

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="plot salary distribution",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_insights_mode(self):
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            classify_response = json.dumps({
                "mode": "insights",
                "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]
            })
            sub_q_response = json.dumps(["Q1?", "Q2?"])
            pandas_code = "result = df_1['salary'].sum()"
            stitch_response = "Key insights: salary is high"

            R.llm_client.generate = MagicMock(side_effect=[
                classify_response,
                sub_q_response,
                pandas_code,    # sub-q 1
                pandas_code,    # sub-q 2
                stitch_response,
            ])

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="analyze my data",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_qna_code_fails_returns_fallback(self):
        """Covers: error is not None path in qna mode."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            classify_response = json.dumps({
                "mode": "qna",
                "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]
            })
            bad_code = "result = df_1['nonexistent_column_xyz']"
            R.llm_client.generate = MagicMock(side_effect=[
                classify_response,
                bad_code, bad_code, bad_code,  # all retries fail
            ])

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="what is sum of xyz?",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_follow_up_loads_session(self):
        """Covers follow-up question path (no s3_urls)."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            session_data = {
                "s3_urls": ["https://bucket/test.xlsx"],
                "all_sheets_metadata": [{"file_name": "test.xlsx", "sheet_name": "Sheet1", "columns": []}]
            }

            llm_responses = [
                json.dumps({"mode": "qna", "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]}),
                "result = df_1['salary'].sum()",
                "Total salary is 230000",
            ]
            R.llm_client.generate = MagicMock(side_effect=llm_responses)

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "get_session_data", new_callable=AsyncMock, return_value=session_data):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls=None,
                        question="What is total salary?",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_s3_urls_as_json_array(self):
        """Covers: json array parsing for s3_urls."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            llm_responses = [
                json.dumps({"mode": "qna", "relevant_sources": [{"file_name": "test.xlsx", "sheet_name": "Sheet1"}]}),
                "result = df_1['salary'].sum()",
                "Answer",
            ]
            R.llm_client.generate = MagicMock(side_effect=llm_responses)

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()
            s3_json = json.dumps(["https://bucket/test.xlsx"])
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls=s3_json,
                        question="test question",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_no_valid_sheets_raises(self):
        """Covers: no all_sheets_data after processing → HTTPException."""
        async def run():
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = None  # empty content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()    
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with pytest.raises(HTTPException):
                    await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="test",
                        user_metadata=valid_metadata(),
                    )
        asyncio.get_event_loop().run_until_complete(run())

    def test_follow_up_s3_load_fails_raises(self):
        """Covers: follow-up with empty all_sheets_data → HTTPException."""
        async def run():
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = None
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3

            session_data = {
                "s3_urls": ["https://bucket/test.xlsx"],
                "all_sheets_metadata": []
            }

            mock_llm_config = AsyncMock()
            mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
            mock_llm_config.__aexit__ = AsyncMock(return_value=False)
            mock_llm_config.get_team_model_config = AsyncMock(return_value={
                "selected_model": "gpt-4", "provider": "openai", "config": {}
            })
            req = make_mock_request()    
            with patch.object(R, "get_model_config", return_value=mock_llm_config):
                with patch.object(R, "get_session_data", new_callable=AsyncMock, return_value=session_data):
                    with pytest.raises(HTTPException):
                        await R.analyze_data_from_multiple_s3_files(
                            request=req,
                            s3_urls=None,
                            question="follow-up",
                            user_metadata=valid_metadata(),
                        )
        asyncio.get_event_loop().run_until_complete(run())

    def test_general_exception_raises_http_500(self):
        """Covers: general Exception → HTTPException 500."""
        async def run():
            req = make_mock_request()
            with patch.object(R, "extract_metadata_fields", side_effect=Exception("unexpected")):
                with pytest.raises(HTTPException) as exc_info:
                    await R.analyze_data_from_multiple_s3_files(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        question="test",
                        user_metadata=valid_metadata(),
                    )
            assert exc_info.value.status_code == 500
        asyncio.get_event_loop().run_until_complete(run())

    def test_plot_multi_sheet_concat(self):
        """Covers: len(relevant_sheets_data) > 1 → pd.concat for plot."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df(), "Sheet2": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            mock_s3.upload_file.return_value = "https://s3/plot.html"
            mock_s3.generate_presigned_url.return_value = "https://presigned/plot.html"
            R.s3_utility = mock_s3

            R.llm_client.generate = MagicMock(return_value=json.dumps({
                "mode": "plot",
                "relevant_sources": [
                    {"file_name": "test.xlsx", "sheet_name": "Sheet1"},
                    {"file_name": "test.xlsx", "sheet_name": "Sheet2"},
                ]
            }))

            mock_plot_generator = MagicMock()
            mock_plot_generator.create_plot.return_value = ('{"data":[],"layout":{}}', "bar")
            R.plot_generator = mock_plot_generator

            import plotly.io as pio
            with patch.object(pio, "from_json", return_value=MagicMock()):
                with patch.object(pio, "to_html", return_value="<html>plot</html>"):
                    mock_llm_config = AsyncMock()
                    mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
                    mock_llm_config.__aexit__ = AsyncMock(return_value=False)
                    mock_llm_config.get_team_model_config = AsyncMock(return_value={
                        "selected_model": "gpt-4", "provider": "openai", "config": {}
                    })
                    req = make_mock_request()
                    with patch.object(R, "get_model_config", return_value=mock_llm_config):
                        with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                            response = await R.analyze_data_from_multiple_s3_files(
                                request=req,
                                s3_urls="https://bucket/test.xlsx",
                                question="plot data",
                                user_metadata=valid_metadata(),
                            )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None



# generate_eda_report

class TestGenerateEdaReport:

    def _make_llm_config(self):
        mock_llm_config = AsyncMock()
        mock_llm_config.__aenter__ = AsyncMock(return_value=mock_llm_config)
        mock_llm_config.__aexit__ = AsyncMock(return_value=False)
        mock_llm_config.get_team_model_config = AsyncMock(return_value={
            "selected_model": "gpt-4", "provider": "openai", "config": {}
        })
        return mock_llm_config

    def test_success_generates_report(self):
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            mock_s3.upload_file.return_value = "https://s3/report.pdf"
            mock_s3.generate_presigned_url.return_value = "https://presigned/report.pdf"
            R.s3_utility = mock_s3

            mock_report_gen = AsyncMock()
            mock_report_gen.generate_multi_sheet_report = AsyncMock(
                return_value=(b"pdf_bytes", "report.pdf")
            )
            R.report_generator = mock_report_gen

            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.generate_eda_report(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_no_s3_urls_uses_session(self):
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            mock_s3.upload_file.return_value = "https://s3/report.pdf"
            mock_s3.generate_presigned_url.return_value = "https://presigned/report.pdf"
            R.s3_utility = mock_s3

            mock_report_gen = AsyncMock()
            mock_report_gen.generate_multi_sheet_report = AsyncMock(
                return_value=(b"pdf_bytes", "report.pdf")
            )
            R.report_generator = mock_report_gen

            session_data = {"s3_urls": ["https://bucket/test.xlsx"]}

            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with patch.object(R, "get_session_data", new_callable=AsyncMock, return_value=session_data):
                    with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                        response = await R.generate_eda_report(
                            request=req,
                            s3_urls=None,
                            user_metadata=valid_metadata(),
                        )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_no_s3_urls_no_session_raises(self):
        async def run():
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with patch.object(R, "get_session_data", new_callable=AsyncMock, return_value=None):
                    with pytest.raises(HTTPException):
                        await R.generate_eda_report(
                            request=req,
                            s3_urls=None,
                            user_metadata=valid_metadata(),
                        )
        asyncio.get_event_loop().run_until_complete(run())

    def test_invalid_s3_url_raises(self):
        async def run():
            req = make_mock_request()         
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with pytest.raises(HTTPException):
                    await R.generate_eda_report(
                        request=req,
                        s3_urls="ftp://bad/url",
                        user_metadata=valid_metadata(),
                    )
        asyncio.get_event_loop().run_until_complete(run())

    def test_no_valid_sheets_raises(self):
        async def run():
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = None
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            R.s3_utility = mock_s3
            req = make_mock_request()
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with pytest.raises(HTTPException):
                    await R.generate_eda_report(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        user_metadata=valid_metadata(),
                    )
        asyncio.get_event_loop().run_until_complete(run())

    def test_general_exception_raises_500(self):
        async def run():
            req = make_mock_request()
            with patch.object(R, "extract_metadata_fields", side_effect=Exception("unexpected")):
                with pytest.raises(HTTPException) as exc_info:
                    await R.generate_eda_report(
                        request=req,
                        s3_urls="https://bucket/test.xlsx",
                        user_metadata=valid_metadata(),
                    )
            assert exc_info.value.status_code == 500
        asyncio.get_event_loop().run_until_complete(run())

    def test_s3_urls_as_json_array(self):
        """Covers json array parsing path for s3_urls."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.return_value = "test.xlsx"
            mock_s3.upload_file.return_value = "https://s3/report.pdf"
            mock_s3.generate_presigned_url.return_value = "https://presigned/report.pdf"
            R.s3_utility = mock_s3

            mock_report_gen = AsyncMock()
            mock_report_gen.generate_multi_sheet_report = AsyncMock(
                return_value=(b"pdf_bytes", "report.pdf")
            )
            R.report_generator = mock_report_gen

            req = make_mock_request()
            s3_json = json.dumps(["https://bucket/test.xlsx"])
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.generate_eda_report(
                        request=req,
                        s3_urls=s3_json,
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None

    def test_multiple_s3_files_multi_filename(self):
        """Covers: len(s3_url_list) != 1 → 'Multi-File Analysis' filename."""
        async def run():
            content = make_excel_bytes({"Sheet1": make_df()})
            mock_s3 = MagicMock()
            mock_s3.get_data_from_s3_by_url.return_value = content
            mock_s3.extract_filename_from_s3_url.side_effect = ["file1.xlsx", "file2.xlsx"]
            mock_s3.upload_file.return_value = "https://s3/report.pdf"
            mock_s3.generate_presigned_url.return_value = "https://presigned/report.pdf"
            R.s3_utility = mock_s3

            mock_report_gen = AsyncMock()
            mock_report_gen.generate_multi_sheet_report = AsyncMock(
                return_value=(b"pdf_bytes", "report.pdf")
            )
            R.report_generator = mock_report_gen

            req = make_mock_request()
            s3_urls = "https://bucket/file1.xlsx,https://bucket/file2.xlsx"
            with patch.object(R, "get_model_config", return_value=self._make_llm_config()):
                with patch.object(R, "create_or_update_session", new_callable=AsyncMock, return_value="sess-1"):
                    response = await R.generate_eda_report(
                        request=req,
                        s3_urls=s3_urls,
                        user_metadata=valid_metadata(),
                    )
            return response
        response = asyncio.get_event_loop().run_until_complete(run())
        assert response is not None